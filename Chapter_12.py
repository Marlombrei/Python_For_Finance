from __future__ import division, print_function
import numpy as np
import numpy.random as npr
import scipy as sp
import scipy.stats as scs
import scipy.optimize as sco
import pandas as pd
from pandas_datareader import data as pdr
from datetime import datetime, date, time
import openpyxl as oxl
from openpyxl import load_workbook
import xlrd, xlwt, xlsxwriter
# import matplotlib as mpl
# import matplotlib.pyplot as plt
# import pandas.core.tools.datetimes as datetools
# import statsmodels.api as sm
# #import statsmodels.formula.api as smf
# import numexpr as ne
# from math import *
# from sklearn.decomposition import KernelPCA



pd.set_option('display.width', 500)
pd.options.display.float_format = '{:,.4f}'.format
pd.options.display.max_rows = 10
np.set_printoptions(linewidth=200)

npr.seed(1000)

path = r'C:\Users\Marlombrei\Documents\Python_For_Finance'

#===============================================================================
# 
# wb = xlwt.Workbook() # In-memory version
# wb.add_sheet(sheetname='first_sheet',   # adding sheet to workbook
#              cell_overwrite_ok=True)
# 
# ws_1 = wb.get_sheet(0) # assigning the first sheet to a variable
# ws_2 = wb.add_sheet(sheetname='second_sheet') # creating a second sheet
# 
# # Adding data to the workbook
# data = np.arange(1,65).reshape((8,8))
# print(data)
# 
# ws_1.write(0,0,100) # write 100 in cell A1
# #wb.save(path+'\workbook.xls')
# 
# wb = xlsxwriter.Workbook(path+'\workbook.xlsx')
# ws_1 = wb.add_worksheet('first_sheet')
# ws_2 = wb.add_worksheet('second_sheet')
# 
# for c in range(data.shape[0]):
#     for r in range(data.shape[1]):
#         ws_1.write(r,c,data[r,c])
#         ws_2.write(r,c,data[r,c])
# 
# wb.close() # it must be explicitly closed
# 
# wb = xlsxwriter.Workbook(path+'\chart.xlsx')
# ws = wb.add_worksheet()
# values = abs(npr.standard_normal(15).cumsum())
# ws.write_column('A1', values)
# chart = wb.add_chart({'type':'line'})
# chart.add_series({'values': '=Sheet1!$A$1:$A$15',
#                   'marker':{'type':'diamond'},})
# ws.insert_chart('C1',chart)
# 
# 
# wb.close()
#===============================================================================

#===============================================================================
data = np.arange(1,65).reshape((8,8))
wb = oxl.Workbook()
ws = wb.create_sheet(title='oxl_sheet', index=0)
for c in range(data.shape[0]):
    for r in range(data.shape[1]):
        ws.cell(row=r+1, column=c+1).value = data[c,r]
 
wb.save(path+'\oxl_book.xlsx')
#===============================================================================


#===============================================================================
# using Pandas
df_1 = pd.read_excel(path+'\workbook.xlsx',
                     'first_sheet',
                     header=None)

df_2 = pd.read_excel(path+'\workbook.xlsx',
                     'second_sheet',
                     header=None)

import string
columns = [string.ascii_uppercase[c] for c in range(data.shape[0])]

df_1.columns = columns
df_2.columns = columns
print(df_1,'\n\n')
print(df_2,'\n\n')
#===============================================================================




































































































